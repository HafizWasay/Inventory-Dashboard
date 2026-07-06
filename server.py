from __future__ import annotations

import csv
import asyncio
import base64
import io
import json
import mimetypes
import os
import re
import shutil
import sys
from datetime import date, datetime
from http.server import ThreadingHTTPServer, SimpleHTTPRequestHandler
from pathlib import Path
from urllib.parse import parse_qs, unquote, urlparse

try:
    from openpyxl import load_workbook
except ImportError:
    raise SystemExit("openpyxl is required. Start this app with run_dashboard.cmd.")

ROOT = Path(__file__).resolve().parent
PUBLIC = ROOT / "public"
IS_VERCEL = bool(os.environ.get("VERCEL"))
SOURCE_DATA = ROOT / "data" / "workbooks"
RUNTIME_ROOT = Path("/tmp/inventory-dashboard") if IS_VERCEL else ROOT / "data"
DATA = RUNTIME_ROOT / "workbooks"
BACKUPS = RUNTIME_ROOT / "backups"
BLOB_PREFIX = "inventory-dashboard/workbooks"
BLOB_AUTH = (
    "oidc"
    if os.environ.get("BLOB_STORE_ID") and os.environ.get("VERCEL_OIDC_TOKEN")
    else "token"
    if os.environ.get("BLOB_READ_WRITE_TOKEN")
    else ""
)
BLOB_ENABLED = IS_VERCEL and bool(BLOB_AUTH)

SOURCES = {
    "assigned": ("Inventory List/Assigned Laptops.xlsx", "laptop", "Assigned"),
    "instock": ("Inventory List/Instock Laptops.xlsx", "laptop", "In stock"),
    "malfunctioned": ("Inventory List/Malfunctioned Laptops.xlsx", "laptop", "Malfunctioned"),
    "snatched": ("Inventory List/Snatched Laptops.xlsx", "laptop", "Snatched"),
    "buyback": ("Inventory List/Buyback Laptops.xlsx", "laptop", "Buyback"),
    "accessories": ("Asset Sheet/Given Headphones & Mouse.xlsx", "accessory", "Assigned"),
    "stock": ("Asset Sheet/Stock Available.xlsx", "stock", "Available"),
    "hiring": ("Upcoming Hiring/Upcoming Hiring Sheet.xlsx", "hire", "Hiring"),
    "extensions": ("Extension List/extensions.xlsx", "extension", "Directory"),
}

STORAGE_ERROR = ""


async def _blob_sync_async():
    from vercel.blob import AsyncBlobClient, BlobNotFoundError
    async with AsyncBlobClient() as client:
        for rel_path, _, _ in SOURCES.values():
            rel = Path(rel_path)
            local_path = DATA / rel
            local_path.parent.mkdir(parents=True, exist_ok=True)
            pathname = f"{BLOB_PREFIX}/{rel.as_posix()}"
            try:
                result = await client.get(pathname, access="private")
            except BlobNotFoundError:
                result = None
            if result and result.status_code == 200:
                local_path.write_bytes(result.content)
            elif (SOURCE_DATA / rel).exists():
                seed = (SOURCE_DATA / rel).read_bytes()
                local_path.write_bytes(seed)
                await client.put(pathname, seed, access="private", overwrite=True)


def ensure_runtime_data():
    global STORAGE_ERROR
    if not IS_VERCEL:
        return
    marker = RUNTIME_ROOT / ".synced"
    if marker.exists():
        return
    DATA.mkdir(parents=True, exist_ok=True)
    for rel_path, _, _ in SOURCES.values():
        source = SOURCE_DATA / rel_path
        target = DATA / rel_path
        if source.exists() and not target.exists():
            target.parent.mkdir(parents=True, exist_ok=True)
            shutil.copy2(source, target)
    if BLOB_ENABLED:
        try:
            asyncio.run(_blob_sync_async())
            STORAGE_ERROR = ""
        except Exception as exc:
            STORAGE_ERROR = f"Blob synchronization failed: {exc}"
    else:
        STORAGE_ERROR = "Vercel Blob credentials are unavailable. Data is read-only and changes cannot persist."
    marker.write_text(datetime.now().isoformat(), encoding="utf-8")


async def _blob_put_async(path, pathname):
    from vercel.blob import AsyncBlobClient
    async with AsyncBlobClient() as client:
        await client.put(
            pathname, path.read_bytes(), access="private", overwrite=True,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )


def persist_workbook(path):
    if not IS_VERCEL:
        return
    if not BLOB_ENABLED:
        raise ValueError("Vercel Blob is not connected. Create a private Blob store before editing deployed data.")
    rel = path.relative_to(DATA).as_posix()
    asyncio.run(_blob_put_async(path, f"{BLOB_PREFIX}/{rel}"))


def persist_backup(path):
    if IS_VERCEL and BLOB_ENABLED:
        asyncio.run(_blob_put_async(path, f"inventory-dashboard/backups/{path.name}"))

IMPORT_FIELDS = {
    "assigned": [
        ("Employee ID", ["employee id", "emp id", "staff id", "id"]),
        ("Employee Name", ["employee name", "user name", "assigned to", "name", "employee"]),
        ("Department ", ["department", "dept", "team", "division"]),
        ("Designation", ["designation", "position", "job title", "role"]),
        ("Make", ["make", "manufacturer", "brand"]),
        ("Eqipment Description", ["equipment description", "eqipment description", "model", "specification", "specs", "device model"]),
        ("Serial #", ["serial", "serial number", "serial no", "service tag"]),
        ("Date Of Purchase", ["date of purchase", "purchase date", "dop"]),
        ("DOP(Year)", ["dop year", "purchase year", "year"]),
        ("Purchase Value", ["purchase value", "cost", "price", "amount"]),
        ("ASSETS TAG", ["asset tag", "assets tag", "asset id", "inventory tag"]),
        ("Date of warrenty expire", ["warranty expiry", "warranty expire", "warranty end"]),
        ("Vendor Name", ["vendor", "vendor name", "supplier"]),
    ],
    "instock": [
        ("Make", ["make", "manufacturer", "brand"]),
        ("Equipment Description", ["equipment description", "eqipment description", "model", "specification", "specs", "device model"]),
        ("Serial #", ["serial", "serial number", "serial no", "service tag"]),
        ("Date Of Purchase", ["date of purchase", "purchase date", "dop"]),
        ("DOP (Year)", ["dop year", "purchase year", "year"]),
        ("Purchase Value", ["purchase value", "cost", "price", "amount"]),
        ("ASSETS TAG", ["asset tag", "assets tag", "asset id", "inventory tag"]),
        ("Vendor", ["vendor", "vendor name", "supplier"]),
    ],
    "accessories": [
        ("Date", ["date", "assigned date", "issue date"]),
        ("Asset Type", ["asset type", "type", "item", "device"]),
        ("Brand/Model", ["brand model", "brand/model", "model", "make"]),
        ("Location", ["location", "office", "site"]),
        ("User/Department", ["user department", "user/department", "assigned to", "employee", "user"]),
        ("Status", ["status", "state"]),
    ],
    "hiring": [
        ("Date", ["date", "start date", "joining date", "date of joining", "doj"]),
        ("Name", ["name", "employee name", "candidate", "new hire"]),
        ("Contact", ["contact", "phone", "mobile", "phone number"]),
        ("Department", ["department", "dept", "team", "division"]),
        ("Designation", ["designation", "position", "job title", "role"]),
        ("Status", ["status", "laptop status", "state"]),
        ("Laptop Asset Tag", ["laptop asset tag", "asset tag", "assets tag"]),
        ("Laptop Serial", ["laptop serial", "serial", "serial number"]),
        ("Configuration Owner", ["configuration owner", "configured by", "it owner"]),
        ("Setup Notes", ["setup notes", "notes", "remarks", "comments"]),
    ],
    "extensions": [
        ("extension", ["extension", "ext", "extension number", "phone extension"]),
        ("name", ["name", "person", "employee", "room", "department"]),
    ],
}


def clean(value):
    if value is None:
        return ""
    if isinstance(value, (datetime, date)):
        return value.isoformat()[:10]
    if isinstance(value, float) and value.is_integer():
        return int(value)
    return str(value).strip() if not isinstance(value, (int, float, bool)) else value


def key(value):
    return re.sub(r"[^a-z0-9]+", "_", str(value).strip().lower()).strip("_")


def import_rows(filename, content_b64):
    try:
        content = base64.b64decode(content_b64, validate=True)
    except Exception as exc:
        raise ValueError("The uploaded file could not be read") from exc
    max_bytes = 3 * 1024 * 1024 if IS_VERCEL else 15 * 1024 * 1024
    if len(content) > max_bytes:
        limit = 3 if IS_VERCEL else 15
        raise ValueError(f"Import files must be {limit} MB or smaller in this environment")
    suffix = Path(filename).suffix.lower()
    if suffix == ".csv":
        text = content.decode("utf-8-sig", errors="replace")
        parsed = list(csv.reader(io.StringIO(text)))
    elif suffix == ".xlsx":
        wb = load_workbook(io.BytesIO(content), data_only=True, read_only=True)
        ws = wb[wb.sheetnames[0]]
        parsed = [[clean(cell) for cell in row] for row in ws.iter_rows(values_only=True)]
        wb.close()
    else:
        raise ValueError("Please upload a CSV or XLSX file")
    if not parsed:
        raise ValueError("The uploaded file is empty")
    headers = [str(clean(v) or f"Column {i+1}") for i, v in enumerate(parsed[0])]
    rows = []
    for values in parsed[1:]:
        record = {headers[i]: clean(v) for i, v in enumerate(values[:len(headers)])}
        if any(v != "" for v in record.values()):
            rows.append(record)
    return headers, rows


def suggested_mapping(source_id, headers):
    suggestions = {}
    normalized = {key(h): h for h in headers}
    for target, aliases in IMPORT_FIELDS[source_id]:
        candidates = [target, *aliases]
        match = next((normalized[key(candidate)] for candidate in candidates if key(candidate) in normalized), "")
        if not match:
            target_tokens = set(key(target).split("_"))
            scored = []
            for header in headers:
                tokens = set(key(header).split("_"))
                score = len(target_tokens & tokens) / max(len(target_tokens | tokens), 1)
                if score >= 0.5:
                    scored.append((score, header))
            match = max(scored, default=(0, ""))[1]
        suggestions[target] = match
    return suggestions


def stock_number(value):
    """Evaluate the stock sheet's simple arithmetic formulas safely."""
    value = clean(value)
    if isinstance(value, (int, float)):
        return value
    text = str(value).strip()
    if text.startswith("="):
        expression = text[1:].replace(" ", "")
        if re.fullmatch(r"\d+(?:[+-]\d+)*", expression):
            return sum(int(part) for part in re.findall(r"[+-]?\d+", expression))
    try:
        return float(text)
    except ValueError:
        return value


def read_xlsx(source_id, rel_path, kind, lifecycle):
    path = DATA / rel_path
    wb = load_workbook(path, data_only=False)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [clean(x) or f"Column {i+1}" for i, x in enumerate(rows[0])]
    records = []
    for excel_row, values in enumerate(rows[1:], start=2):
        raw = {headers[i]: clean(v) for i, v in enumerate(values) if i < len(headers) and clean(v) != ""}
        if not raw:
            continue
        if source_id == "stock":
            item = clean(values[0] if len(values) > 0 else "")
            units = stock_number(values[1] if len(values) > 1 else "")
            if item and str(item).lower() != "items" and units != "":
                records.append({
                    "id": f"{source_id}:{excel_row}", "source": source_id, "row": excel_row,
                    "kind": kind, "lifecycle": lifecycle, "name": item, "assetType": item,
                    "quantity": units, "status": "Low stock" if isinstance(units, (int, float)) and units < 5 else "Healthy",
                    "location": "", "department": "", "serial": "", "assetTag": "", "details": raw,
                })
            continue
        if source_id == "extensions":
            extension = raw.get("extension", "")
            person = raw.get("name", "")
            records.append({
                "id": f"{source_id}:{excel_row}", "source": source_id, "row": excel_row,
                "kind": kind, "lifecycle": lifecycle, "name": person or f"Extension {extension}",
                "assetType": "Extension", "status": "Directory", "department": "", "location": "",
                "serial": str(extension), "assetTag": "", "description": "", "quantity": 1, "details": raw,
            })
            continue
        name = raw.get("Employee Name") or raw.get("Name") or raw.get("User/Department") or raw.get("Make") or "Unassigned"
        asset_type = raw.get("Asset Type") or ("Laptop" if kind == "laptop" else kind.title())
        status = raw.get("Laptop Status") or raw.get("Status") or lifecycle
        if kind == "hire" and not raw.get("Laptop Status"):
            status = "Pending" if str(raw.get("Status", "")).strip().lower() == "pending" else "Done"
        department = raw.get("Department") or raw.get("Department ") or ""
        location = raw.get("Location") or ""
        serial = raw.get("Serial #") or ""
        asset_tag = raw.get("ASSETS TAG") or ""
        description = raw.get("Equipment Description") or raw.get("Eqipment Description") or raw.get("Brand/Model") or ""
        records.append({
            "id": f"{source_id}:{excel_row}", "source": source_id, "row": excel_row,
            "kind": kind, "lifecycle": lifecycle, "name": name, "assetType": asset_type,
            "status": status, "department": department, "location": location, "serial": serial,
            "assetTag": asset_tag, "description": description, "quantity": 1, "details": raw,
        })
    return records


def all_data():
    records = []
    for source_id, (rel_path, kind, lifecycle) in SOURCES.items():
        records.extend(read_xlsx(source_id, rel_path, kind, lifecycle))
    return records


def backup(path):
    BACKUPS.mkdir(parents=True, exist_ok=True)
    stamp = datetime.now().strftime("%Y%m%d-%H%M%S-%f")
    backup_path = BACKUPS / f"{path.stem}-{stamp}{path.suffix}"
    shutil.copy2(path, backup_path)
    persist_backup(backup_path)


def update_record(payload):
    source_id = payload.get("source")
    row = int(payload.get("row", 0))
    updates = payload.get("updates") or {}
    if source_id not in SOURCES or row < 2 or not isinstance(updates, dict):
        raise ValueError("Invalid update request")
    rel_path = SOURCES[source_id][0]
    path = DATA / rel_path
    backup(path)
    wb = load_workbook(path)
    ws = wb[wb.sheetnames[0]]
    headers = {clean(cell.value): idx for idx, cell in enumerate(ws[1], start=1) if clean(cell.value)}
    for field, value in updates.items():
        field = str(field).strip()
        if not field or field.startswith("_"):
            continue
        if source_id == "stock" and field == "Units":
            ws.cell(row, 2, value if value != "" else None)
            continue
        if field not in headers:
            col = ws.max_column + 1
            ws.cell(1, col, field)
            headers[field] = col
        ws.cell(row, headers[field], value if value != "" else None)
    temp = path.with_suffix(".tmp.xlsx")
    wb.save(temp)
    wb.close()
    os.replace(temp, path)
    persist_workbook(path)
    return {"ok": True, "record": next((r for r in read_xlsx(source_id, rel_path, SOURCES[source_id][1], SOURCES[source_id][2]) if r["row"] == row), None)}


def add_record(payload):
    source_id = payload.get("source")
    values = payload.get("values") or {}
    if source_id not in {"assigned", "instock", "accessories", "hiring", "stock", "extensions"} or not isinstance(values, dict):
        raise ValueError("Invalid add request")
    rel_path, kind, lifecycle = SOURCES[source_id]
    path = DATA / rel_path
    backup(path)
    wb = load_workbook(path)
    ws = wb[wb.sheetnames[0]]
    if source_id == "stock":
        next_row = ws.max_row + 1
        ws.cell(next_row, 1, values.get("Item", ""))
        ws.cell(next_row, 2, values.get("Units", 0))
        temp = path.with_suffix(".tmp.xlsx")
        wb.save(temp)
        wb.close()
        os.replace(temp, path)
        persist_workbook(path)
        record = next((r for r in read_xlsx(source_id, rel_path, kind, lifecycle) if r["row"] == next_row), None)
        return {"ok": True, "record": record}
    headers = {clean(cell.value): idx for idx, cell in enumerate(ws[1], start=1) if clean(cell.value)}
    next_row = ws.max_row + 1
    if "S#" in headers:
        serials = [ws.cell(row, headers["S#"]).value for row in range(2, ws.max_row + 1)]
        numeric = [int(v) for v in serials if isinstance(v, (int, float))]
        ws.cell(next_row, headers["S#"], max(numeric, default=0) + 1)
    for field, value in values.items():
        field = str(field).strip()
        if not field or field.startswith("_") or value == "":
            continue
        if field not in headers:
            col = ws.max_column + 1
            ws.cell(1, col, field)
            headers[field] = col
        ws.cell(next_row, headers[field], value)
    temp = path.with_suffix(".tmp.xlsx")
    wb.save(temp)
    wb.close()
    os.replace(temp, path)
    persist_workbook(path)
    record = next((r for r in read_xlsx(source_id, rel_path, kind, lifecycle) if r["row"] == next_row), None)
    return {"ok": True, "record": record}


def append_values(ws, values):
    headers = {clean(cell.value): idx for idx, cell in enumerate(ws[1], start=1) if clean(cell.value)}
    next_row = ws.max_row + 1
    if "S#" in headers:
        serials = [ws.cell(row, headers["S#"]).value for row in range(2, ws.max_row + 1)]
        numeric = [int(v) for v in serials if isinstance(v, (int, float))]
        ws.cell(next_row, headers["S#"], max(numeric, default=0) + 1)
    for field, value in values.items():
        if value in ("", None):
            continue
        if field not in headers:
            col = ws.max_column + 1
            ws.cell(1, col, field)
            headers[field] = col
        ws.cell(next_row, headers[field], value)
    return next_row


def delete_record(payload):
    source_id = payload.get("source")
    row = int(payload.get("row", 0))
    if source_id not in SOURCES or row < 2:
        raise ValueError("Invalid delete request")
    path = DATA / SOURCES[source_id][0]
    wb = load_workbook(path)
    ws = wb[wb.sheetnames[0]]
    if row > ws.max_row:
        wb.close()
        raise ValueError("Record no longer exists")
    backup(path)
    ws.delete_rows(row, 1)
    temp = path.with_suffix(".tmp.xlsx")
    wb.save(temp)
    wb.close()
    os.replace(temp, path)
    persist_workbook(path)
    return {"ok": True, "deleted": True}


def move_record(payload):
    source_id = payload.get("source")
    destination = payload.get("destination")
    row = int(payload.get("row", 0))
    extras = payload.get("values") or {}
    allowed = {"assigned", "instock", "malfunctioned", "buyback"}
    if source_id not in allowed or destination not in allowed or source_id == destination or row < 2:
        raise ValueError("Invalid move request")
    source_path = DATA / SOURCES[source_id][0]
    destination_path = DATA / SOURCES[destination][0]
    source_wb = load_workbook(source_path)
    destination_wb = load_workbook(destination_path)
    source_ws = source_wb[source_wb.sheetnames[0]]
    destination_ws = destination_wb[destination_wb.sheetnames[0]]
    if row > source_ws.max_row:
        source_wb.close(); destination_wb.close()
        raise ValueError("Record no longer exists")
    source_headers = [clean(cell.value) for cell in source_ws[1]]
    raw_values = [clean(cell.value) for cell in source_ws[row]]
    raw = {source_headers[i]: raw_values[i] for i in range(min(len(source_headers), len(raw_values))) if source_headers[i]}
    common = {
        "Make": raw.get("Make", ""),
        "description": raw.get("Equipment Description") or raw.get("Eqipment Description") or "",
        "Serial #": raw.get("Serial #", ""),
        "Date Of Purchase": raw.get("Date Of Purchase", ""),
        "year": raw.get("DOP(Year)") or raw.get("DOP (Year)") or "",
        "Purchase Value": raw.get("Purchase Value", ""),
        "ASSETS TAG": raw.get("ASSETS TAG", ""),
        "vendor": raw.get("Vendor Name") or raw.get("Vendor") or "",
        "Employee Name": raw.get("Employee Name") or raw.get("Name") or "",
        "Department": raw.get("Department ") or raw.get("Department") or "",
        "Designation": raw.get("Designation") or raw.get("Position") or "",
    }
    if destination == "instock":
        values = {
            "Make": common["Make"], "Equipment Description": common["description"], "Serial #": common["Serial #"],
            "Date Of Purchase": common["Date Of Purchase"], "DOP (Year)": common["year"],
            "Purchase Value": common["Purchase Value"], "ASSETS TAG": common["ASSETS TAG"], "Vendor": common["vendor"],
        }
    elif destination == "assigned":
        values = {
            "Employee ID": extras.get("Employee ID", raw.get("Employee ID", "")),
            "Employee Name": extras.get("Employee Name") or common["Employee Name"],
            "Department ": extras.get("Department") or common["Department"],
            "Designation": extras.get("Designation") or common["Designation"],
            "Make": common["Make"], "Eqipment Description": common["description"], "Serial #": common["Serial #"],
            "Date Of Purchase": common["Date Of Purchase"], "DOP(Year)": common["year"],
            "Purchase Value": common["Purchase Value"], "ASSETS TAG": common["ASSETS TAG"], "Vendor Name": common["vendor"],
        }
        if not values["Employee Name"]:
            source_wb.close(); destination_wb.close()
            raise ValueError("Employee name is required when assigning a laptop")
    elif destination == "malfunctioned":
        values = {
            "Employee Name": extras.get("Employee Name") or common["Employee Name"], "Make": common["Make"],
            "Eqipment Description": common["description"], "Serial #": common["Serial #"],
            "Date Of Purchase": common["Date Of Purchase"], "Purchase Value": common["Purchase Value"],
            "ASSETS TAG": common["ASSETS TAG"], "Vendor": common["vendor"], "Issue": extras.get("Issue", ""),
        }
    else:
        values = {
            "ID": extras.get("Employee ID", raw.get("Employee ID", "")),
            "Name": extras.get("Employee Name") or common["Employee Name"], "Department": extras.get("Department") or common["Department"],
            "Position": extras.get("Designation") or common["Designation"], "Eqipment Description": common["description"],
            "Serial #": common["Serial #"], "Date Of Purchase": common["Date Of Purchase"],
            "BuyBack Date ": extras.get("BuyBack Date", ""), "Purchase Value": common["Purchase Value"], "ASSETS TAG": common["ASSETS TAG"],
        }
    backup(source_path)
    backup(destination_path)
    destination_row = append_values(destination_ws, values)
    source_ws.delete_rows(row, 1)
    source_temp = source_path.with_suffix(".tmp.xlsx")
    destination_temp = destination_path.with_suffix(".tmp.xlsx")
    destination_wb.save(destination_temp)
    source_wb.save(source_temp)
    destination_wb.close(); source_wb.close()
    os.replace(destination_temp, destination_path)
    os.replace(source_temp, source_path)
    persist_workbook(destination_path)
    persist_workbook(source_path)
    return {"ok": True, "moved": True, "destination": destination, "row": destination_row}


def import_preview(payload):
    source_id = payload.get("source")
    if source_id not in IMPORT_FIELDS:
        raise ValueError("Choose a valid destination dataset")
    headers, rows = import_rows(payload.get("filename", ""), payload.get("content", ""))
    return {
        "headers": headers,
        "rowCount": len(rows),
        "sample": rows[:3],
        "fields": [field for field, _ in IMPORT_FIELDS[source_id]],
        "suggestions": suggested_mapping(source_id, headers),
    }


def import_commit(payload):
    source_id = payload.get("source")
    mapping = payload.get("mapping") or {}
    mode = payload.get("mode", "append")
    if source_id not in IMPORT_FIELDS or mode not in {"append", "replace"}:
        raise ValueError("Invalid import request")
    headers, rows = import_rows(payload.get("filename", ""), payload.get("content", ""))
    valid_headers = set(headers)
    mapping = {target: source for target, source in mapping.items() if source in valid_headers}
    if not mapping:
        raise ValueError("Map at least one column before importing")
    rel_path = SOURCES[source_id][0]
    path = DATA / rel_path
    backup(path)
    wb = load_workbook(path)
    ws = wb[wb.sheetnames[0]]
    target_headers = {clean(cell.value): idx for idx, cell in enumerate(ws[1], start=1) if clean(cell.value)}
    for target in mapping:
        if target not in target_headers:
            col = ws.max_column + 1
            ws.cell(1, col, target)
            target_headers[target] = col
    if mode == "replace" and ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)
    start_row = ws.max_row + 1
    serial_col = target_headers.get("S#")
    for offset, record in enumerate(rows):
        row_num = start_row + offset
        if serial_col:
            ws.cell(row_num, serial_col, offset + 1 if mode == "replace" else row_num - 1)
        for target, source in mapping.items():
            value = record.get(source, "")
            if value != "":
                ws.cell(row_num, target_headers[target], value)
    temp = path.with_suffix(".tmp.xlsx")
    wb.save(temp)
    wb.close()
    os.replace(temp, path)
    persist_workbook(path)
    return {"ok": True, "imported": len(rows), "mode": mode, "dataset": source_id}


def export_csv(source_id):
    output = io.StringIO(newline="")
    if source_id == "all":
        fields = ["Dataset", "Lifecycle", "Asset Type", "Name / User", "Department", "Location", "Status", "Serial Number", "Asset Tag", "Description", "Quantity"]
        writer = csv.DictWriter(output, fieldnames=fields)
        writer.writeheader()
        for record in all_data():
            writer.writerow({
                "Dataset": record["source"], "Lifecycle": record["lifecycle"], "Asset Type": record["assetType"],
                "Name / User": record["name"], "Department": record["department"], "Location": record["location"],
                "Status": record["status"], "Serial Number": record["serial"], "Asset Tag": record["assetTag"],
                "Description": record.get("description", ""), "Quantity": record["quantity"],
            })
        filename = "inventory-dashboard-all.csv"
    elif source_id in SOURCES:
        records = [r for r in all_data() if r["source"] == source_id]
        fields = []
        for record in records:
            for field in record.get("details", {}):
                if field not in fields and not field.startswith("Unnamed:"):
                    fields.append(field)
        if source_id == "stock":
            fields = ["Item", "Units", "Status"]
            rows = [{"Item": r["name"], "Units": r["quantity"], "Status": r["status"]} for r in records]
        else:
            rows = [{field: record.get("details", {}).get(field, "") for field in fields} for record in records]
        writer = csv.DictWriter(output, fieldnames=fields)
        writer.writeheader()
        writer.writerows(rows)
        filename = f"{source_id}-export.csv"
    else:
        raise ValueError("Unknown export dataset")
    return filename, output.getvalue().encode("utf-8-sig")


class Handler(SimpleHTTPRequestHandler):
    def api_path(self):
        parsed = urlparse(self.path)
        routed = parse_qs(parsed.query).get("route", [""])[0]
        if routed:
            return f"/api/{routed.lstrip('/')}"
        return parsed.path

    def translate_path(self, path):
        requested = unquote(urlparse(path).path).lstrip("/") or "index.html"
        target = (PUBLIC / requested).resolve()
        if PUBLIC.resolve() not in target.parents and target != PUBLIC.resolve():
            return str(PUBLIC / "index.html")
        return str(target)

    def send_json(self, body, status=200):
        encoded = json.dumps(body, ensure_ascii=False).encode("utf-8")
        self.send_response(status)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Content-Length", str(len(encoded)))
        self.end_headers()
        self.wfile.write(encoded)

    def do_GET(self):
        parsed_url = urlparse(self.path)
        route = self.api_path()
        if route.startswith("/api/"):
            ensure_runtime_data()
        if route == "/api/data":
            try:
                records = all_data()
                self.send_json({
                    "records": records, "updatedAt": datetime.now().isoformat(timespec="seconds"),
                    "storage": {"persistent": not IS_VERCEL or BLOB_ENABLED, "warning": STORAGE_ERROR},
                })
            except Exception as exc:
                self.send_json({"error": str(exc)}, 500)
            return
        if route == "/api/health":
            self.send_json({
                "ok": True, "environment": "vercel" if IS_VERCEL else "local",
                "blobConnected": BLOB_ENABLED, "blobAuth": BLOB_AUTH or None,
                "storageWarning": STORAGE_ERROR,
                "sources": len(SOURCES),
            })
            return
        if route == "/api/export":
            try:
                source_id = parse_qs(parsed_url.query).get("source", ["all"])[0]
                filename, content = export_csv(source_id)
                self.send_response(200)
                self.send_header("Content-Type", "text/csv; charset=utf-8")
                self.send_header("Content-Disposition", f'attachment; filename="{filename}"')
                self.send_header("Content-Length", str(len(content)))
                self.end_headers()
                self.wfile.write(content)
            except Exception as exc:
                self.send_json({"error": str(exc)}, 400)
            return
        super().do_GET()

    def do_POST(self):
        route = self.api_path()
        ensure_runtime_data()
        if route not in {"/api/update", "/api/add", "/api/delete", "/api/move", "/api/import/preview", "/api/import/commit"}:
            self.send_json({"error": "Not found"}, 404)
            return
        try:
            length = int(self.headers.get("Content-Length", 0))
            if IS_VERCEL and length > 4_400_000:
                self.send_json({"error": "This import is too large for Vercel. Use a file smaller than 3 MB."}, 413)
                return
            payload = json.loads(self.rfile.read(length) or b"{}")
            if route == "/api/add":
                result = add_record(payload)
            elif route == "/api/delete":
                result = delete_record(payload)
            elif route == "/api/move":
                result = move_record(payload)
            elif route == "/api/import/preview":
                result = import_preview(payload)
            elif route == "/api/import/commit":
                result = import_commit(payload)
            else:
                result = update_record(payload)
            self.send_json(result)
        except Exception as exc:
            self.send_json({"error": str(exc)}, 400)

    def log_message(self, fmt, *args):
        print(f"[dashboard] {fmt % args}")


if __name__ == "__main__":
    port = int(os.environ.get("PORT", "8787"))
    server = ThreadingHTTPServer(("127.0.0.1", port), Handler)
    print(f"Inventory dashboard ready at http://127.0.0.1:{port}")
    try:
        server.serve_forever()
    except KeyboardInterrupt:
        print("\nDashboard stopped.")
